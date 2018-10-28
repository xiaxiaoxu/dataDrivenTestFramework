# encoding=utf-8

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
import time


class parseExcel(object):
    def __init__(self, excelPath):
        self.excelPath = excelPath
        self.workbook = load_workbook(excelPath)  # 加载excel
        self.sheet = self.workbook.active  # 获取第一个sheet
        self.font = Font(color=None)
        self.colorDict = {"red": 'FFFF3030', "green": 'FF008B00'}

    # 设置当前要操作的sheet对象，使用index来获取相应的sheet
    def set_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.get_sheet_names()[sheet_index]

        self.sheet = self.workbook.get_sheet_by_name(sheet_name)
        return self.sheet

    # 获取当前默认sheet的名字
    def get_default_sheet(self):
        return self.sheet.title

    # 设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet
    def set_sheet_by_name(self, sheet_name):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        self.sheet = sheet
        return self.sheet

    # 获取默认sheet中最大的行数
    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取默认 sheet 的最大列数
    def get_max_col_no(self):
        return self.sheet.max_column

    # 获取默认sheet的最小（起始）行号
    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet的最小（起始）列号
    def get_min_col_no(self):
        return self.sheet.min_column

    # 获取默认 sheet 的所有行对象，
    def get_all_rows(self):
        return list(self.sheet.iter_rows())
        # return list(self.rows)也可以

    # 获取默认sheet中的所有列对象
    def get_all_cols(self):
        return list(self.sheet.iter_cols())
        # return list(self.sheet.columns)也可以

    # 从默认sheet中获取某一列，第一列从0开始
    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    # 从默认sheet中获取某一行，第一行从0开始
    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    # 从默认sheet中，通过行号和列号获取指定的单元格，注意行号和列号从1开始
    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    # 从默认sheet中，通过行号和列号获取指定的单元格中的内容，注意行号和列号从1开始
    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入指定内容，注意行号和列号从1开始
    # 调用此方法的时候，excel不要处于打开状态
    def write_cell_content(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期，注意行号和列号从1开始
    # 调用此方法的时候，excel不要处于打开状态
    def write_cell_current_time(self, row_no, col_no):
        time1 = time.strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.cell(row=row_no, column=col_no).value = str(time1)
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=col_no).value

    def save_excel_file(self):
        self.workbook.save(self.excelPath)


if __name__ == '__main__':
    p = parseExcel(u'D:\\testdata.xlsx')
    print u"获取默认行：", p.get_default_sheet()

    print u"设置sheet索引为1", p.set_sheet_by_index(1)
    print u"获取默认行：", p.get_default_sheet()
    print u"设置sheet索引为0", p.set_sheet_by_index(0)
    print u"获取默认行：", p.get_default_sheet()
    # for i in range(3,6):
    # for j in range(3,6):
    # p.write_cell_content(i,j,str((i,j)))

    print u"最大行数：", p.get_max_row_no()
    print u"最大列数：", p.get_max_col_no()
    print u"最小起始行数：", p.get_min_row_no()
    print u"最小起始列数：", p.get_min_col_no()
    print u"所有行对象：", p.get_all_rows()
    print u"所有列对象：", p.get_all_cols()
    print u"获取某一列(2)：", p.get_single_col(2)
    print u"获取某一行(4)：", p.get_single_row(4)
    print u"取得行号和列号(2,2)单元格：", p.get_cell(2, 2)
    print u"取得行号和列号单元格的内容(2,2)", p.get_cell_content(2, 2)
    print u"行号和列号写入内容(11,11)：'xiaxiaoxu'", p.write_cell_content(11, 11, 'xiaxiaoxu')  #
    print u"行号和列号写入当前日期(13,13)：", p.write_cell_current_time(13, 13)

